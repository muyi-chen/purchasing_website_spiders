import requests, os
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

class Client():
    # 初始化
    def __init__(self, org_name, tender_name, start_date, end_date):
        self.org_name = org_name
        self.tender_name = tender_name
        self.start_date = start_date
        self.end_date = end_date
        self.url = 'https://web.pcc.gov.tw'
        self.headers = {
            'User-Agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36'
        }


    # 搜尋資料
    def search(self):
        params = {
            'pageSize': '50',
            'firstSearch': 'true',
            'searchType': 'basic',
            'isBinding': 'N',
            'isLogIn': 'N',
            'level_1': 'on',
            'orgName': self.org_name,
            'orgId': '',
            'tenderName': self.tender_name,
            'tenderId': '',
            'tenderType': 'TENDER_DECLARATION',
            'tenderWay': 'TENDER_WAY_ALL_DECLARATION',
            'dateType': 'isDate',
            'tenderStartDate': self.start_date,
            'tenderEndDate': self.end_date,
            'radProctrgCate': '',
            'policyAdvocacy': ''
        }

        search_url = self.url + '/prkms/tender/common/basic/readTenderBasic'
        response = requests.get(search_url, params=params, headers=self.headers)

        html_content = response.text

        soup = BeautifulSoup(html_content, 'lxml')

        tender_partial_info = self.get_tender_partial_info(soup)

        tender_info = self.get_tender_info(tender_partial_info)

        self.save_to_excel(tender_info)

        return tender_info


    # 取得標案資訊
    def get_tender_partial_info(self, soup):
        result = []

        org_names_and_tender_numbers = self.get_org_names_and_tender_numbers(soup)
        for i in org_names_and_tender_numbers:
            org_number = self.get_org_number(i['org_name'])
            result.append({
                'org_name': i['org_name'],
                'tender_number': i['tender_number'],
                'org_number': org_number
            })
        
        return result


    # 取得機關完整名稱, 標案案號
    def get_org_names_and_tender_numbers(self, soup):
        result = []

        table = soup.find('table', id='tpam')

        rows = table.find_all('tr') if table else []

        for row in rows:
            cells = row.find_all('td')
            org_name = ''
            tender_number = ''

            if len(cells) > 1:
                org_name = cells[1].text.strip()

            if len(cells) > 2:
                tender_number = cells[2].text.strip().replace(' (更正公告)', '')

            if org_name != '' and tender_number != '':
                result.append({
                    'org_name': org_name,
                    'tender_number': tender_number
                })

        return result


    # 取得機關代碼
    def get_org_number(self, org_name):
        params = {
            'keyWord': org_name
        }
        response = requests.post('https://web.pcc.gov.tw/prkms/tender/common/orgName/search', params=params)

        soup = BeautifulSoup(response.text, 'lxml')

        element = soup.find(lambda tag: tag.name == 'td' and tag.text.strip() == org_name)
        org_number = ''
        if element:
            org_number = element.find_previous_sibling('td').text.strip()

        return org_number


    # 取得標案詳細資訊
    def get_tender_info(self, tender_info):
        result = []
        for i in tender_info:
            params = {
                'unit_id': i['org_number'],
                'job_number': i['tender_number']
            }

            response = requests.get('https://pcc.g0v.ronny.tw/api/tender', params=params)
            response_json = response.json()

            details = {
                '機關名稱': response_json['records'][0]['brief']['title'],
                '標案名稱': response_json['records'][0]['brief']['title'],
                '標案案號': response_json['records'][0]['job_number'],
                '預算金額': response_json['records'][0]['detail']['採購資料:預算金額'],
                '公告日': response_json['records'][0]['detail']['招標資料:公告日'],
                '招標方式': response_json['records'][0]['detail']['招標資料:招標方式'],
                '截止投標': response_json['records'][0]['detail']['領投開標:截止投標'],
                '開標時間': response_json['records'][0]['detail']['領投開標:開標時間'],
                '履約期限': response_json['records'][0]['detail']['其他:履約期限']
            }

            result.append(details)

        return result
    
    
    # 保存Excel
    def save_to_excel(self, data):
        filename = 'tender_info.xlsx'

        if os.path.exists(filename):
            wb = load_workbook(filename)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Tender Info'

            titles = ['機關名稱', '標案名稱', '標案案號', '預算金額', '公告日', '招標方式', '截止投標', '開標時間', '履約期限']
            ws.append(titles)

        tender_numbers = [row[2].value for row in ws.iter_rows(min_row=2, max_col=3)]
        for item in data:
            if item['標案案號'] not in tender_numbers:
                row = [
                    item['機關名稱'],
                    item['標案名稱'],
                    item['標案案號'],
                    item['預算金額'],
                    item['公告日'],
                    item['招標方式'],
                    item['截止投標'],
                    item['開標時間'],
                    item['履約期限'],
                ]
                ws.append(row)

        wb.save(filename)